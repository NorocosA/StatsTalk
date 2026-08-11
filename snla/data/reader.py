"""
StatsTalk Data Reader — sole authority on variable dict structure.

Reads .sav (SPSS) and .csv files, extracting variable metadata
for the LLM pipeline. Raw data values are NEVER exposed to cloud APIs.

CONTRACT: This module is the **single writer** of the variable dict
structure consumed by all downstream modules (planner, _pipeline,
mcp_server, validator, etc.). Every variable dict produced by
extract_metadata() has the shape:
    {"name": str, "type": str, "label": str, "value_labels": dict | None}

If you need to add/change variable dict fields, do it here.  Do NOT
create variable dicts from scratch in other modules — trust this
contract instead.
"""

import os
from pathlib import Path
from zipfile import BadZipFile

import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.cell import range_boundaries
from openpyxl.utils.exceptions import InvalidFileException

try:
    import pyreadstat
except ImportError:
    pyreadstat = None  # type: ignore[assignment]


MAX_EXCEL_EFFECTIVE_CELLS = 5_000_000


class ExcelImportError(ValueError):
    """Actionable error raised while inspecting or loading an Excel workbook."""

    def __init__(self, code: str, message: str):
        self.code = code
        super().__init__(f"{code}: {message}")


def _worksheet_dimensions(worksheet) -> tuple[int, int]:
    try:
        dimension = worksheet.calculate_dimension()
    except ValueError:
        raise ExcelImportError(
            "EXCEL_DIMENSIONS_UNAVAILABLE",
            "The worksheet does not declare safe dimensions and cannot be imported.",
        ) from None
    min_column, min_row, max_column, max_row = range_boundaries(dimension)
    if (min_column, min_row, max_column, max_row) == (1, 1, 1, 1):
        first = next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True), (None,))
        if not first or all(value is None for value in first):
            return 0, 0
    return max_row, max_column


def inspect_xlsx(file_path: str | Path) -> dict:
    """Return sheet names and dimensions without exposing headers or cell values."""

    path = Path(file_path)
    if path.suffix.lower() != ".xlsx":
        raise ExcelImportError("EXCEL_UNSUPPORTED_TYPE", "Only .xlsx workbooks are supported.")
    if not path.is_file():
        raise FileNotFoundError(f"File not found: {path}")

    try:
        workbook = load_workbook(
            path,
            read_only=True,
            data_only=True,
            keep_links=False,
        )
    except (BadZipFile, InvalidFileException, KeyError, OSError, ValueError):
        raise ExcelImportError(
            "EXCEL_CORRUPT", "The workbook is damaged or is not a valid .xlsx file."
        ) from None

    try:
        sheets = []
        total_cells = 0
        for worksheet in workbook.worksheets:
            row_count, column_count = _worksheet_dimensions(worksheet)
            effective_cells = row_count * column_count
            total_cells += effective_cells
            if total_cells > MAX_EXCEL_EFFECTIVE_CELLS:
                raise ExcelImportError(
                    "EXCEL_TOO_MANY_CELLS",
                    f"Workbook exceeds the {MAX_EXCEL_EFFECTIVE_CELLS:,}-cell limit.",
                )
            sheets.append(
                {
                    "name": worksheet.title,
                    "row_count": row_count,
                    "column_count": column_count,
                    "effective_cells": effective_cells,
                }
            )
        return {
            "filename": path.name,
            "format": "xlsx",
            "sheets": sheets,
            "total_effective_cells": total_cells,
        }
    finally:
        workbook.close()


def read_xlsx_sheet(file_path: str | Path, worksheet_name: str) -> tuple["pd.DataFrame", dict]:
    """Load one explicitly selected worksheet using cached formula values only."""

    path = Path(file_path)
    structure = inspect_xlsx(path)
    if worksheet_name not in {sheet["name"] for sheet in structure["sheets"]}:
        raise ExcelImportError(
            "EXCEL_WORKSHEET_NOT_FOUND", "Choose one of the workbook's listed worksheets."
        )

    workbook = load_workbook(path, read_only=True, data_only=True, keep_links=False)
    try:
        worksheet = workbook[worksheet_name]
        rows = worksheet.iter_rows(values_only=True)
        header = list(next(rows, ()))
        if not header or all(value is None for value in header):
            raise ExcelImportError("EXCEL_EMPTY_SHEET", "The selected worksheet is empty.")

        second_row = list(next(rows, ()))
        empty_header = [
            index for index, value in enumerate(header) if value is None or not str(value).strip()
        ]
        if empty_header:
            second_row_looks_like_header = second_row and all(
                isinstance(value, str) and value.strip() for value in second_row[: len(header)]
            )
            if second_row_looks_like_header:
                raise ExcelImportError(
                    "EXCEL_MULTIPLE_HEADER_ROWS",
                    "Use a single unmerged header row before importing this worksheet.",
                )
            raise ExcelImportError(
                "EXCEL_EMPTY_HEADER", "Every imported column must have a non-empty header."
            )

        headers = [str(value).strip() for value in header]
        normalized = [value.casefold() for value in headers]
        if len(normalized) != len(set(normalized)):
            raise ExcelImportError("EXCEL_DUPLICATE_HEADERS", "Column headers must be unique.")

        data_rows = []
        if second_row:
            data_rows.append(second_row[: len(headers)])
        data_rows.extend(list(row[: len(headers)]) for row in rows)
        while data_rows and all(value is None for value in data_rows[-1]):
            data_rows.pop()
        dataframe = pd.DataFrame(data_rows, columns=headers)
        metadata = {
            "filename": path.name,
            "format": "xlsx",
            "worksheet": worksheet_name,
            "row_count": len(dataframe),
            "column_count": len(headers),
            "file_path": str(path.resolve()),
            "file_label": None,
        }
        return dataframe, metadata
    finally:
        workbook.close()


def read_xlsx_and_extract(file_path: str | Path, worksheet_name: str) -> dict:
    """Load one worksheet and return the canonical dataset metadata contract."""

    dataframe, metadata = read_xlsx_sheet(file_path, worksheet_name)
    return extract_metadata(dataframe, metadata)


def read_sav(file_path: str) -> tuple["pd.DataFrame", dict]:
    """
    Read an SPSS .sav file and extract metadata.

    Args:
        file_path: Path to .sav file

    Returns:
        (dataframe, metadata) tuple
        metadata: {
            "filename": str,
            "format": "sav",
            "row_count": int,
            "column_count": int,
            "file_path": str,
            "file_label": str | None,
            # Internal keys (consumed by extract_metadata):
            "_column_names": list[str],
            "_column_labels": list[str | None],
            "_variable_value_labels": dict,
        }

    Raises:
        FileNotFoundError: If file doesn't exist
        ImportError: If pyreadstat is not installed
        ValueError: If file is not a valid .sav file
    """
    if pyreadstat is None:
        raise ImportError("pyreadstat is not installed. Install it with: pip install pyreadstat")

    if not os.path.exists(file_path):
        raise FileNotFoundError(f"File not found: {file_path}")

    df, meta = pyreadstat.read_sav(file_path)

    column_names = list(getattr(meta, "column_names", df.columns))
    column_labels = getattr(meta, "column_labels", None)
    value_labels = getattr(meta, "variable_value_labels", {})

    metadata = {
        "filename": os.path.basename(file_path),
        "format": "sav",
        "row_count": len(df),
        "column_count": len(df.columns),
        "file_path": os.path.abspath(file_path),
        "file_label": getattr(meta, "file_label", None),
        # Internal: passed through to extract_metadata
        "_column_names": column_names,
        "_column_labels": column_labels or [None] * len(column_names),
        "_variable_value_labels": value_labels,
    }

    return df, metadata


def read_csv(file_path: str, encoding: str = "utf-8") -> tuple["pd.DataFrame", dict]:
    """
    Read a CSV file and extract basic metadata.

    Args:
        file_path: Path to .csv file
        encoding: File encoding (default utf-8, try gbk for Chinese CSV files)

    Returns:
        (dataframe, metadata) tuple
        metadata: {
            "filename": str,
            "format": "csv",
            "row_count": int,
            "column_count": int,
            "file_path": str,
            "file_label": None,
        }

    Raises:
        FileNotFoundError: If file doesn't exist
        ValueError: If encoding fails after both utf-8 and gbk attempts
    """
    if not os.path.exists(file_path):
        raise FileNotFoundError(f"File not found: {file_path}")

    try:
        df = pd.read_csv(file_path, encoding=encoding)
    except UnicodeDecodeError:
        # Chinese CSV files often use GBK encoding
        try:
            df = pd.read_csv(file_path, encoding="gbk")
        except UnicodeDecodeError as exc:
            raise ValueError(f"Failed to read CSV with utf-8 or gbk encoding: {file_path}") from exc

    metadata = {
        "filename": os.path.basename(file_path),
        "format": "csv",
        "row_count": len(df),
        "column_count": len(df.columns),
        "file_path": os.path.abspath(file_path),
        "file_label": None,
    }

    return df, metadata


def extract_metadata(df: "pd.DataFrame", meta: dict) -> dict:
    """
    Extract unified variable metadata from a dataframe and file metadata.

    Combines pandas dtypes with any pyreadstat variable metadata.
    The output is the canonical variable list format used throughout SNLA.

    Works for both .sav files (pyreadstat metadata available) and .csv files
    (metadata derived purely from pandas dtypes and column names).

    Args:
        df: Pandas DataFrame (from read_sav or read_csv)
        meta: File-level metadata dict (from read_sav or read_csv)

    Returns:
        Unified metadata dict:
        {
            "filename": str,
            "format": "sav" | "csv",
            "row_count": int,
            "column_count": int,
            "file_path": str,
            "file_label": str | None,
            "variables": [
                {
                    "name": str,
                    "type": "Numeric" | "String" | "Date",
                    "label": str,
                    "value_labels": dict | None,
                },
                ...
            ]
        }
    """
    variables = []

    # Pull pyreadstat data from internal keys if present (SAV files).
    # For CSV files these fall back to pandas-derived values.
    column_names = meta.get("_column_names", list(df.columns))
    column_labels = meta.get("_column_labels", [""] * len(df.columns))
    value_labels_dict = meta.get("_variable_value_labels", {})

    for i, col_name in enumerate(column_names):
        if col_name not in df.columns:
            continue

        # Determine type from pandas dtype
        dtype = df[col_name].dtype
        if pd.api.types.is_numeric_dtype(dtype):
            var_type = "Numeric"
        elif pd.api.types.is_datetime64_any_dtype(dtype):
            var_type = "Date"
        else:
            var_type = "String"

        # Get label (handle NaN, None, and non-string values gracefully)
        label = ""
        if i < len(column_labels):
            raw_label = column_labels[i]
            if isinstance(raw_label, str):
                label = raw_label
            else:
                try:
                    # NaN check (NaN != NaN)
                    if raw_label is not None and not (
                        isinstance(raw_label, float) and raw_label != raw_label
                    ):
                        label = str(raw_label)
                except Exception:
                    label = ""

        # Get value labels (if any) — ensure JSON-compatible keys
        value_labels = value_labels_dict.get(col_name)
        if value_labels is not None:
            value_labels = {str(k): v for k, v in value_labels.items()}

        variables.append(
            {
                "name": col_name,
                "type": var_type,
                "label": label,
                "value_labels": value_labels,
            }
        )

    result = dict(meta)
    # Drop internal keys before returning
    result.pop("_column_names", None)
    result.pop("_column_labels", None)
    result.pop("_variable_value_labels", None)
    result["variables"] = variables
    return result


def read_and_extract(file_path: str) -> dict:
    """
    Convenience function: read a file and extract metadata in one call.

    Auto-detects format by file extension (.sav or .csv).

    Args:
        file_path: Path to .sav or .csv file

    Returns:
        Unified metadata dict with variables list

    Raises:
        FileNotFoundError: If file doesn't exist
        ValueError: If file format is not supported
    """
    ext = os.path.splitext(file_path)[1].lower()

    if ext == ".sav":
        df, file_meta = read_sav(file_path)
    elif ext == ".csv":
        df, file_meta = read_csv(file_path)
    else:
        raise ValueError(f"Unsupported file format: {ext}. Expected .sav or .csv")

    return extract_metadata(df, file_meta)
