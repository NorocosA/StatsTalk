from __future__ import annotations

import asyncio
import io
from types import SimpleNamespace

import pytest
from openpyxl import Workbook

from snla.data.reader import ExcelImportError, inspect_xlsx, read_xlsx_sheet


def _workbook_bytes(sheets: dict[str, list[list[object]]]) -> bytes:
    workbook = Workbook()
    workbook.remove(workbook.active)
    for name, rows in sheets.items():
        worksheet = workbook.create_sheet(name)
        for row in rows:
            worksheet.append(row)
    stream = io.BytesIO()
    workbook.save(stream)
    return stream.getvalue()


def test_inspect_xlsx_returns_only_safe_structure(tmp_path):
    path = tmp_path / "private.xlsx"
    path.write_bytes(
        _workbook_bytes(
            {
                "Survey": [["patient_name", "score"], ["Alice", 91]],
                "Archive": [["secret"], ["do-not-leak"]],
            }
        )
    )

    result = inspect_xlsx(path)

    assert result == {
        "filename": "private.xlsx",
        "format": "xlsx",
        "sheets": [
            {"name": "Survey", "row_count": 2, "column_count": 2, "effective_cells": 4},
            {"name": "Archive", "row_count": 2, "column_count": 1, "effective_cells": 2},
        ],
        "total_effective_cells": 6,
    }
    assert "Alice" not in repr(result)
    assert "patient_name" not in repr(result)


def test_read_xlsx_sheet_loads_only_explicit_sheet_with_cached_values(tmp_path):
    path = tmp_path / "survey.xlsx"
    path.write_bytes(
        _workbook_bytes(
            {
                "Selected": [["group", "score"], ["A", 3], ["B", 5]],
                "Ignored": [["private"], ["never-read"]],
            }
        )
    )

    dataframe, meta = read_xlsx_sheet(path, "Selected")

    assert dataframe.to_dict(orient="records") == [
        {"group": "A", "score": 3},
        {"group": "B", "score": 5},
    ]
    assert meta["worksheet"] == "Selected"
    assert meta["row_count"] == 2
    assert meta["column_count"] == 2


def test_read_xlsx_sheet_never_exposes_formula_source(tmp_path):
    path = tmp_path / "formula.xlsx"
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.append(["score"])
    worksheet["A2"] = "=1+1"
    workbook.save(path)

    dataframe, _ = read_xlsx_sheet(path, worksheet.title)

    assert dataframe.empty
    assert "=1+1" not in repr(dataframe)


@pytest.mark.parametrize(
    ("rows", "code"),
    [
        ([], "EXCEL_EMPTY_SHEET"),
        ([["score", "score"], [1, 2]], "EXCEL_DUPLICATE_HEADERS"),
        ([["score", None], [1, 2]], "EXCEL_EMPTY_HEADER"),
    ],
)
def test_read_xlsx_sheet_rejects_bad_headers(tmp_path, rows, code):
    path = tmp_path / "bad.xlsx"
    path.write_bytes(_workbook_bytes({"Sheet1": rows}))

    with pytest.raises(ExcelImportError, match=code):
        read_xlsx_sheet(path, "Sheet1")


def test_read_xlsx_sheet_rejects_merged_multirow_header(tmp_path):
    path = tmp_path / "multi-header.xlsx"
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Survey"
    worksheet.merge_cells("A1:B1")
    worksheet["A1"] = "Scores"
    worksheet.append(["pre", "post"])
    workbook.save(path)

    with pytest.raises(ExcelImportError, match="EXCEL_MULTIPLE_HEADER_ROWS"):
        read_xlsx_sheet(path, "Survey")


def test_inspect_xlsx_rejects_corrupt_and_oversized_workbooks(tmp_path):
    corrupt = tmp_path / "corrupt.xlsx"
    corrupt.write_bytes(b"not-a-zip")
    with pytest.raises(ExcelImportError, match="EXCEL_CORRUPT"):
        inspect_xlsx(corrupt)

    oversized = tmp_path / "oversized.xlsx"
    workbook = Workbook()
    workbook.active.cell(row=1_000_001, column=5, value=1)
    workbook.save(oversized)
    with pytest.raises(ExcelImportError, match="EXCEL_TOO_MANY_CELLS"):
        inspect_xlsx(oversized)


@pytest.fixture
def excel_client(tmp_path, monkeypatch):
    from snla import config
    from snla.data.retention import DatasetRetention
    from snla.ui import server
    from snla.ui.security import loopback_security

    class Provider:
        def protect(self, plaintext):
            return plaintext

        def unprotect(self, ciphertext):
            return ciphertext

    monkeypatch.setattr(config, "SESSION_RESTORE_ENABLED", False)
    monkeypatch.setattr(
        server,
        "dataset_retention",
        DatasetRetention(
            reference_path=tmp_path / "restore.bin",
            workspace_root=tmp_path / "workspaces",
            provider=Provider(),
            restore_enabled=lambda: False,
        ),
    )
    server.session.reset()
    server.app.config["TESTING"] = True
    with server.app.test_client() as client:
        token = loopback_security.begin_launch("http://127.0.0.1:43125")
        bootstrap = client.post("/api/bootstrap", json={"bootstrap_token": token})
        client.environ_base["HTTP_AUTHORIZATION"] = (
            f"Bearer {bootstrap.get_json()['session_token']}"
        )
        yield client
    server.session.reset()


def test_http_excel_requires_explicit_worksheet_before_loading(excel_client):
    contents = _workbook_bytes(
        {"Wave 1": [["group", "score"], ["A", 1]], "Wave 2": [["group"], ["B"]]}
    )

    upload = excel_client.post(
        "/api/upload",
        data={
            "file": (
                io.BytesIO(contents),
                "survey.xlsx",
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )

    assert upload.status_code == 200
    body = upload.get_json()
    assert body["requires_worksheet_selection"] is True
    assert [sheet["name"] for sheet in body["sheets"]] == ["Wave 1", "Wave 2"]
    assert "variables" not in body

    selected = excel_client.post("/api/select-worksheet", json={"worksheet": "Wave 1"})
    assert selected.status_code == 200
    selected_body = selected.get_json()
    assert selected_body["worksheet"] == "Wave 1"
    assert [item["name"] for item in selected_body["variables"]] == ["group", "score"]


def test_http_excel_returns_actionable_error_for_corrupt_workbook(excel_client):
    response = excel_client.post(
        "/api/upload",
        data={
            "file": (
                io.BytesIO(b"not-an-xlsx"),
                "broken.xlsx",
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )

    assert response.status_code == 422
    assert response.get_json()["code"] == "EXCEL_CORRUPT"


def test_local_excel_rejects_100mb_before_workbook_inspection(monkeypatch):
    from snla.ui import server

    class OversizedWorkbook:
        suffix = ".xlsx"

        def stat(self):
            return SimpleNamespace(st_size=server.MAX_EXCEL_UPLOAD_SIZE + 1)

    monkeypatch.setattr(
        server,
        "inspect_xlsx",
        lambda path: pytest.fail("oversized workbook must not be inspected"),
    )

    with pytest.raises(ExcelImportError, match="EXCEL_FILE_TOO_LARGE"):
        server._load_dataset_path(OversizedWorkbook(), filename="large.xlsx")


@pytest.mark.parametrize("filename", ["legacy.xls", "macro.xlsm", "binary.xlsb"])
def test_http_rejects_unsupported_excel_types(excel_client, filename):
    response = excel_client.post("/api/upload", data={"file": (io.BytesIO(b"data"), filename)})

    assert response.status_code == 400


class _Context:
    session_id = "excel-test"

    async def info(self, message):
        return None


def test_mcp_excel_requires_explicit_worksheet(tmp_path, monkeypatch):
    from snla import mcp_server

    source = tmp_path / "survey.xlsx"
    source.write_bytes(_workbook_bytes({"Data": [["group", "score"], ["A", 1]]}))
    monkeypatch.setattr(mcp_server, "_upload_dir", tmp_path / "mcp")
    mcp_server._session_states.clear()

    upload = asyncio.run(mcp_server.snla_upload(_Context(), str(source)))
    assert upload["requires_worksheet_selection"] is True
    assert "variables" not in upload

    selected = asyncio.run(mcp_server.snla_select_worksheet(_Context(), "Data"))
    assert selected["ok"] is True
    assert [item["name"] for item in selected["variables"]] == ["group", "score"]


def test_mcp_rejects_corrupt_and_unsupported_excel(tmp_path, monkeypatch):
    from snla import mcp_server

    monkeypatch.setattr(mcp_server, "_upload_dir", tmp_path / "mcp")
    mcp_server._session_states.clear()
    corrupt = tmp_path / "broken.xlsx"
    corrupt.write_bytes(b"not-an-xlsx")
    legacy = tmp_path / "legacy.xls"
    legacy.write_bytes(b"legacy")

    corrupt_result = asyncio.run(mcp_server.snla_upload(_Context(), str(corrupt)))
    legacy_result = asyncio.run(mcp_server.snla_upload(_Context(), str(legacy)))

    assert corrupt_result["error"]["code"] == "EXCEL_CORRUPT"
    assert legacy_result["error"]["code"] == "UNSUPPORTED_FILE_TYPE"
